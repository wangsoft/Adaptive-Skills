from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from adaptive_skills.catalog import Catalog
from adaptive_skills.config import Settings
from adaptive_skills.inventory import InventoryBridge
from adaptive_skills.scanner import CatalogScanner
from adaptive_skills.sources import SourceManager

from tests.helpers import commit_all, init_repo, write_skill
from tests.test_cli import run_cli


@unittest.skipUnless(
    importlib.util.find_spec("openpyxl"), "openpyxl is an optional dependency"
)
class InventoryTests(unittest.TestCase):
    def test_existing_columns_import_and_safe_export(self) -> None:
        import openpyxl

        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw)
            library = root / "library"
            library.mkdir()
            repo = init_repo(library / "source")
            write_skill(repo, "excel-skill", "=UNTRUSTED() spreadsheet workflow")
            commit_all(repo)
            settings = Settings.load(library)
            source = SourceManager(settings).register(repo, name="source")
            CatalogScanner(settings).scan(source["id"])

            existing = root / "existing.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "技能总表"
            sheet.append(
                [
                    "评分",
                    "评分来源",
                    "一级分类",
                    "细分类",
                    "技能名",
                    "解决的问题",
                    "应用场景",
                    "备注 / 注意事项",
                    "相对路径",
                ]
            )
            sheet.append(
                [
                    8.2,
                    "人工",
                    "数据",
                    "表格",
                    "excel-skill",
                    "整理数据",
                    "报表",
                    "检查公式",
                    "source/skills/excel-skill/SKILL.md",
                ]
            )
            workbook.create_sheet("评分说明")["A1"] = "preserve me"
            workbook.save(existing)
            workbook.close()

            bridge = InventoryBridge(settings)
            result = bridge.import_xlsx(existing)
            self.assertEqual(result["imported"], 1)
            skill = Catalog(settings).get_skill("excel-skill")
            self.assertEqual(skill["category_l1"], "数据")

            output = root / "updated.xlsx"
            bridge.export_xlsx(output, template=existing)
            exported = openpyxl.load_workbook(output, data_only=False)
            self.assertIn("评分说明", exported.sheetnames)
            self.assertEqual(exported["评分说明"]["A1"].value, "preserve me")
            headers = [cell.value for cell in exported["技能总表"][1]]
            description_column = headers.index("原始 description（触发词见此）") + 1
            self.assertTrue(
                str(exported["技能总表"].cell(2, description_column).value).startswith(
                    "'="
                )
            )
            self.assertGreaterEqual(len(exported["技能总表"].conditional_formatting), 3)
            self.assertEqual(exported["技能总表"]["B2"].number_format, "0.0")
            exported.close()

            cli_output = root / "cli-export.xlsx"
            cli_export = run_cli(
                library, "inventory", "export", "--output", str(cli_output)
            )
            self.assertEqual(cli_export["output"], str(cli_output.resolve()))
            self.assertTrue(cli_output.is_file())
