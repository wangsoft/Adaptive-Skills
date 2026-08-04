from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any

from .catalog import Catalog
from .config import Settings
from .database import path_is_within, utc_now
from .errors import ConflictError, NotFoundError, ValidationError
from .sources import SourceManager


INVENTORY_SHEET = "技能总表"
SOURCES_SHEET = "Sources"
EXPORT_HEADERS = [
    "Skill ID",
    "评分",
    "评分来源",
    "一级分类",
    "细分类",
    "技能名",
    "解决的问题",
    "应用场景",
    "备注 / 注意事项",
    "标签",
    "来源仓库",
    "来源 URL",
    "相对路径",
    "绝对路径",
    "有效",
    "安全等级",
    "Commit",
    "内容哈希",
    "SKILL.md 行数",
    "附带文件数",
    "原始 description（触发词见此）",
    "更新时间",
]


def _openpyxl():
    try:
        import openpyxl
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValidationError(
            "Excel support is optional. Install it with: pip install 'adaptive-skills[excel]'"
        ) from exc
    return (
        openpyxl,
        Alignment,
        Font,
        PatternFill,
        get_column_letter,
        CellIsRule,
        FormulaRule,
    )


def _normalized(value: Any) -> str:
    return str(value or "").strip().replace("\\", "/")


def _excel_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    # Prevent untrusted repository text from becoming an Excel formula.
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


class InventoryBridge:
    def __init__(self, settings: Settings):
        self.settings = settings
        self.catalog = Catalog(settings)
        self.sources = SourceManager(settings)

    def import_xlsx(
        self, workbook_path: str | Path, *, sheet: str = INVENTORY_SHEET
    ) -> dict[str, Any]:
        openpyxl, *_ = _openpyxl()
        path = Path(workbook_path).expanduser().resolve()
        if not path.is_file():
            raise NotFoundError(f"Workbook does not exist: {path}")
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in workbook.sheetnames:
            workbook.close()
            raise NotFoundError(f"Workbook has no sheet named {sheet!r}")
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            workbook.close()
            raise ValidationError(f"Inventory sheet is empty: {sheet}")
        headers = {
            _normalized(value): index
            for index, value in enumerate(raw_headers)
            if value is not None
        }

        skills = self.catalog.list_skills(include_inactive=True)
        by_id = {skill["id"]: skill for skill in skills}
        by_path: dict[str, list[dict[str, Any]]] = {}
        by_name: dict[str, list[dict[str, Any]]] = {}
        for skill in skills:
            keys = {
                _normalized(skill["skill_md_path"]),
                _normalized(
                    f"{Path(skill['source_path']).name}/{skill['rel_path']}/SKILL.md"
                ),
                _normalized(f"{skill['rel_path']}/SKILL.md"),
            }
            absolute = Path(skill["skill_md_path"])
            if path_is_within(absolute, self.settings.library):
                keys.add(absolute.relative_to(self.settings.library).as_posix())
            for key in keys:
                by_path.setdefault(key, []).append(skill)
            by_name.setdefault(skill["name"], []).append(skill)

        imported = 0
        unmatched: list[dict[str, Any]] = []
        ambiguous: list[dict[str, Any]] = []
        for row_number, row in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue

            def get(name: str) -> Any:
                return (
                    row[headers[name]]
                    if name in headers and headers[name] < len(row)
                    else None
                )

            skill: dict[str, Any] | None = None
            stable_id = _normalized(get("Skill ID"))
            relative = _normalized(get("相对路径"))
            absolute = _normalized(get("绝对路径"))
            name = _normalized(get("技能名"))
            if stable_id and stable_id in by_id:
                skill = by_id[stable_id]
            else:
                matches: list[dict[str, Any]] = []
                for key in (absolute, relative):
                    if key:
                        matches.extend(by_path.get(key, []))
                matches = list({item["id"]: item for item in matches}.values())
                if len(matches) == 1:
                    skill = matches[0]
                elif len(matches) > 1:
                    ambiguous.append(
                        {"row": row_number, "name": name, "path": relative or absolute}
                    )
                    continue
                elif name and len(by_name.get(name, [])) == 1:
                    skill = by_name[name][0]
            if skill is None:
                unmatched.append(
                    {"row": row_number, "name": name, "path": relative or absolute}
                )
                continue

            score = get("评分")
            if score not in (None, ""):
                try:
                    score = float(score)
                except (TypeError, ValueError):
                    score = None
            tags = get("标签")
            self.catalog.annotate(
                skill["id"],
                category_l1=get("一级分类"),
                category_l2=get("细分类") or get("二级分类"),
                problem=get("解决的问题"),
                use_case=get("应用场景"),
                score=score,
                score_source=get("评分来源"),
                notes=get("备注 / 注意事项") or get("备注"),
                tags=tags or [],
            )
            imported += 1
        workbook.close()
        return {
            "workbook": str(path),
            "sheet": sheet,
            "imported": imported,
            "unmatched": unmatched,
            "ambiguous": ambiguous,
        }

    def export_xlsx(
        self,
        output_path: str | Path,
        *,
        template: str | Path | None = None,
    ) -> dict[str, Any]:
        (
            openpyxl,
            Alignment,
            Font,
            PatternFill,
            get_column_letter,
            CellIsRule,
            FormulaRule,
        ) = _openpyxl()
        output = Path(output_path).expanduser().resolve()
        template_path = Path(template).expanduser().resolve() if template else None
        if template_path and not template_path.is_file():
            raise NotFoundError(f"Template workbook does not exist: {template_path}")
        if output.exists() and output.is_dir():
            raise ConflictError(f"Output is a directory: {output}")
        workbook = (
            openpyxl.load_workbook(template_path)
            if template_path
            else openpyxl.Workbook()
        )
        if not template_path and workbook.active:
            workbook.remove(workbook.active)

        inventory_index = (
            workbook.sheetnames.index(INVENTORY_SHEET)
            if INVENTORY_SHEET in workbook.sheetnames
            else 0
        )
        if INVENTORY_SHEET in workbook.sheetnames:
            workbook.remove(workbook[INVENTORY_SHEET])
        worksheet = workbook.create_sheet(INVENTORY_SHEET, inventory_index)
        worksheet.append(EXPORT_HEADERS)
        skills = self.catalog.list_skills(include_inactive=True)
        for skill in skills:
            absolute = Path(skill["skill_md_path"])
            relative = (
                absolute.relative_to(self.settings.library).as_posix()
                if path_is_within(absolute, self.settings.library)
                else f"{Path(skill['source_path']).name}/{skill['rel_path']}/SKILL.md"
            )
            worksheet.append(
                [
                    skill["id"],
                    skill.get("score"),
                    skill.get("score_source"),
                    skill.get("category_l1"),
                    skill.get("category_l2"),
                    skill["name"],
                    skill.get("problem"),
                    skill.get("use_case"),
                    skill.get("notes"),
                    ", ".join(skill.get("tags", [])),
                    skill["source_name"],
                    skill.get("source_url"),
                    relative,
                    skill["skill_md_path"],
                    "是" if skill["valid"] else "否",
                    skill["audit_severity"],
                    skill.get("head_sha"),
                    skill["content_hash"],
                    skill["line_count"],
                    skill["file_count"],
                    skill["description"],
                    skill["updated_at"],
                ]
            )

        sources_index = (
            workbook.sheetnames.index(SOURCES_SHEET)
            if SOURCES_SHEET in workbook.sheetnames
            else len(workbook.sheetnames)
        )
        if SOURCES_SHEET in workbook.sheetnames:
            workbook.remove(workbook[SOURCES_SHEET])
        sources_sheet = workbook.create_sheet(SOURCES_SHEET, sources_index)
        source_headers = [
            "Source ID",
            "名称",
            "URL",
            "本地路径",
            "跟踪分支",
            "Commit",
            "状态",
            "最近扫描",
            "更新时间",
        ]
        sources_sheet.append(source_headers)
        for source in self.sources.list():
            sources_sheet.append(
                [
                    source["id"],
                    source["name"],
                    source["url"],
                    source["local_path"],
                    source["tracked_ref"],
                    source["head_sha"],
                    source["status"],
                    source["last_scanned_at"],
                    source["updated_at"],
                ]
            )

        for sheet_object in (worksheet, sources_sheet):
            for cell in sheet_object[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            for row in sheet_object.iter_rows(min_row=2):
                for cell in row:
                    cell.value = _excel_text(cell.value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            sheet_object.freeze_panes = "A2"
            sheet_object.auto_filter.ref = sheet_object.dimensions
            for index, column in enumerate(sheet_object.columns, start=1):
                max_length = max(
                    (len(str(cell.value or "")) for cell in column), default=8
                )
                sheet_object.column_dimensions[get_column_letter(index)].width = min(
                    max(max_length + 2, 10), 48
                )

        worksheet.row_dimensions[1].height = 30
        if worksheet.max_row > 1:
            score_range = f"B2:B{worksheet.max_row}"
            score_fills = [
                ("greaterThanOrEqual", "8.5", "63BE7B"),
                ("greaterThanOrEqual", "7.5", "A9D18E"),
                ("greaterThanOrEqual", "6.5", "FFE699"),
                ("greaterThanOrEqual", "5.5", "F4B183"),
                ("lessThan", "5.5", "F8696B"),
            ]
            for operator, threshold, color in score_fills:
                worksheet.conditional_formatting.add(
                    score_range,
                    CellIsRule(
                        operator=operator,
                        formula=[threshold],
                        fill=PatternFill("solid", fgColor=color),
                        stopIfTrue=True,
                    ),
                )
            worksheet["B2"].number_format = "0.0"
            for cell in worksheet["B"][2:]:
                cell.number_format = "0.0"

            valid_range = f"O2:O{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                valid_range,
                FormulaRule(
                    formula=['$O2="否"'],
                    fill=PatternFill("solid", fgColor="F8696B"),
                ),
            )
            risk_range = f"P2:P{worksheet.max_row}"
            for value, color in (
                ("critical", "C00000"),
                ("high", "F8696B"),
                ("medium", "FFE699"),
                ("low", "C6E0B4"),
            ):
                worksheet.conditional_formatting.add(
                    risk_range,
                    FormulaRule(
                        formula=[f'$P2="{value}"'],
                        fill=PatternFill("solid", fgColor=color),
                        font=(
                            Font(color="FFFFFF", bold=True)
                            if value == "critical"
                            else None
                        ),
                    ),
                )

        output.parent.mkdir(parents=True, exist_ok=True)
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{output.stem}.", suffix=".xlsx", dir=output.parent
        )
        os.close(descriptor)
        temporary = Path(temporary_name)
        try:
            workbook.save(temporary)
            os.replace(temporary, output)
        finally:
            workbook.close()
            if temporary.exists():
                temporary.unlink()
        return {
            "output": str(output),
            "template": str(template_path) if template_path else None,
            "skills": len(skills),
            "sources": len(self.sources.list()),
            "generated_at": utc_now(),
        }
